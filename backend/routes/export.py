from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from ..models.budget import DepartmentBudget
from ..models.department import Department
from ..models.salary import Salary
from ..models.employee import Employee
from ..models.budget_adjustment import BudgetAdjustment
from .. import db
from sqlalchemy import and_, func, distinct, or_
from datetime import datetime
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font

export_bp = Blueprint('export', __name__)

@export_bp.route('/budget/excel', methods=['GET'])
@jwt_required()
def export_budget_excel():
    """导出预算报表为Excel"""
    try:
        year = request.args.get('year', datetime.now().year, type=int)
        month = request.args.get('month', datetime.now().month, type=int)
        
        # 获取预算数据
        budget_data = db.session.query(
            Department.name.label('department'),
            DepartmentBudget.total_budget,
            DepartmentBudget.salary_budget,
            DepartmentBudget.overtime_budget,
            DepartmentBudget.bonus_budget,
            DepartmentBudget.other_budget,
            func.sum(Salary.net_salary).label('actual_expense'),
            func.count(distinct(Employee.id)).label('employee_count')
        ).join(
            DepartmentBudget,
            and_(
                Department.id == DepartmentBudget.department_id,
                DepartmentBudget.year == year,
                DepartmentBudget.month == month
            )
        ).join(
            Employee,
            Employee.department_id == Department.id
        ).join(
            Salary,
            and_(
                Salary.employee_id == Employee.id,
                Salary.year == year,
                Salary.month == month
            ),
            isouter=True
        ).group_by(
            Department.name,
            DepartmentBudget.total_budget,
            DepartmentBudget.salary_budget,
            DepartmentBudget.overtime_budget,
            DepartmentBudget.bonus_budget,
            DepartmentBudget.other_budget
        ).all()
        
        # 转换为DataFrame
        df = pd.DataFrame([{
            '部门': row[0],
            '总预算': float(row[1]),
            '工资预算': float(row[2]),
            '加班预算': float(row[3]),
            '奖金预算': float(row[4]),
            '其他预算': float(row[5]),
            '实际支出': float(row[6]) if row[6] else 0,
            '执行率': f"{(float(row[6])/float(row[1])*100 if row[6] else 0):.1f}%",
            '员工人数': row[7],
            '人均预算': f"{float(row[1])/row[7]:.2f}" if row[7] > 0 else 0
        } for row in budget_data])
        
        # 创建Excel文件
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='预算报表', index=False)
            
            # 获取工作表
            worksheet = writer.sheets['预算报表']
            
            # 设置列宽
            for idx, col in enumerate(df.columns):
                worksheet.column_dimensions[chr(65 + idx)].width = 15
                
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'budget_report_{year}_{month}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500 

@export_bp.route('/budget/structure/excel', methods=['GET'])
@jwt_required()
def export_budget_structure_excel():
    """导出部门费用结构分析报表"""
    try:
        year = request.args.get('year', datetime.now().year, type=int)
        month = request.args.get('month', datetime.now().month, type=int)
        
        # 获取部门费用结构数据
        structure_data = db.session.query(
            Department.name.label('department'),
            func.sum(Salary.base_salary).label('base_salary'),
            func.sum(Salary.overtime_pay).label('overtime_pay'),
            func.sum(Salary.bonus).label('bonus'),
            func.sum(Salary.social_security).label('social_security'),
            func.sum(Salary.net_salary).label('total_expense'),
            func.count(distinct(Employee.id)).label('employee_count'),
            DepartmentBudget.total_budget
        ).join(
            DepartmentBudget,
            and_(
                Department.id == DepartmentBudget.department_id,
                DepartmentBudget.year == year,
                DepartmentBudget.month == month
            )
        ).join(
            Employee,
            Employee.department_id == Department.id
        ).join(
            Salary,
            and_(
                Salary.employee_id == Employee.id,
                Salary.year == year,
                Salary.month == month
            )
        ).group_by(
            Department.name,
            DepartmentBudget.total_budget
        ).all()
        
        # 转换为DataFrame
        df = pd.DataFrame([{
            '部门': row[0],
            '基本工资': float(row[1]),
            '加班工资': float(row[2]),
            '奖金': float(row[3]),
            '社保费用': float(row[4]),
            '总支出': float(row[5]),
            '员工人数': row[6],
            '部门预算': float(row[7]),
            '基本工资占比': f"{(float(row[1])/float(row[5])*100):.1f}%",
            '加班工资占比': f"{(float(row[2])/float(row[5])*100):.1f}%",
            '奖金占比': f"{(float(row[3])/float(row[5])*100):.1f}%",
            '社保占比': f"{(float(row[4])/float(row[5])*100):.1f}%",
            '人均支出': f"{float(row[5])/row[6]:.2f}",
            '预算执行率': f"{(float(row[5])/float(row[7])*100):.1f}%"
        } for row in structure_data])
        
        # 创建Excel文件
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='费用结构分析', index=False)
            
            # 获取工作表
            worksheet = writer.sheets['费用结构分析']
            
            # 设置列宽
            for idx, col in enumerate(df.columns):
                worksheet.column_dimensions[chr(65 + idx)].width = 15
                
            # 设置数字格式
            for row in worksheet.iter_rows(min_row=2):
                for cell in row[1:8]:  # 金额列
                    cell.number_format = '#,##0.00'
                
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'expense_structure_{year}_{month}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/budget/adjustments/excel', methods=['GET'])
@jwt_required()
def export_budget_adjustments_excel():
    """导出预算调整历史记录报表"""
    try:
        year = request.args.get('year', datetime.now().year, type=int)
        month = request.args.get('month', datetime.now().month, type=int)
        department_id = request.args.get('department_id', type=int)
        
        # 构建基础查询
        query = db.session.query(
            Department.name.label('department'),
            DepartmentBudget.year,
            DepartmentBudget.month,
            BudgetAdjustment.created_at,
            BudgetAdjustment.original_amount,
            BudgetAdjustment.adjusted_amount,
            BudgetAdjustment.adjustment_type,
            BudgetAdjustment.reason,
            BudgetAdjustment.notes
        ).join(
            DepartmentBudget,
            BudgetAdjustment.budget_id == DepartmentBudget.id
        ).join(
            Department,
            DepartmentBudget.department_id == Department.id
        ).filter(
            DepartmentBudget.year == year
        )
        
        # 应用筛选条件
        if month:
            query = query.filter(DepartmentBudget.month == month)
        if department_id:
            query = query.filter(Department.id == department_id)
            
        # 获取数据并排序
        adjustment_data = query.order_by(
            Department.name,
            DepartmentBudget.year,
            DepartmentBudget.month,
            BudgetAdjustment.created_at
        ).all()
        
        # 转换为DataFrame
        df = pd.DataFrame([{
            '部门': row[0],
            '年份': row[1],
            '月份': row[2],
            '调整时间': row[3].strftime('%Y-%m-%d %H:%M:%S'),
            '原预算金额': float(row[4]),
            '调整后金额': float(row[5]),
            '调整金额': float(row[5] - row[4]),
            '调整比例': f"{((float(row[5]) - float(row[4])) / float(row[4]) * 100):.1f}%",
            '调整类型': row[6],
            '调整原因': row[7],
            '备注': row[8] or ''
        } for row in adjustment_data])
        
        # 创建Excel文件
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='预算调整历史', index=False)
            
            # 获取工作表
            worksheet = writer.sheets['预算调整历史']
            
            # 设置列宽
            for idx, col in enumerate(df.columns):
                worksheet.column_dimensions[chr(65 + idx)].width = 18
                
            # 设置数字格式
            for row in worksheet.iter_rows(min_row=2):
                for cell in row[4:7]:  # 金额列
                    cell.number_format = '#,##0.00'
                    
            # 设置调整时间列格式
            for row in worksheet.iter_rows(min_row=2):
                row[3].number_format = 'yyyy-mm-dd hh:mm:ss'
                
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'budget_adjustments_{year}{"_"+str(month) if month else ""}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/budget/exceptions/excel', methods=['GET'])
@jwt_required()
def export_budget_exceptions_excel():
    """导出预算执行异常报告"""
    try:
        year = request.args.get('year', datetime.now().year, type=int)
        month = request.args.get('month', datetime.now().month, type=int)
        threshold = request.args.get('threshold', 90, type=int)  # 默认阈值90%
        
        # 获取预算执行异常数据
        exception_data = db.session.query(
            Department.name.label('department'),
            DepartmentBudget.total_budget,
            DepartmentBudget.salary_budget,
            DepartmentBudget.overtime_budget,
            DepartmentBudget.bonus_budget,
            DepartmentBudget.other_budget,
            func.sum(Salary.net_salary).label('actual_expense'),
            func.sum(Salary.base_salary).label('actual_salary'),
            func.sum(Salary.overtime_pay).label('actual_overtime'),
            func.sum(Salary.bonus).label('actual_bonus'),
            func.count(distinct(Employee.id)).label('employee_count')
        ).join(
            DepartmentBudget,
            and_(
                Department.id == DepartmentBudget.department_id,
                DepartmentBudget.year == year,
                DepartmentBudget.month == month,
                DepartmentBudget.status == 'approved'  # 只查询已审批的预算
            )
        ).join(
            Employee,
            Employee.department_id == Department.id
        ).join(
            Salary,
            and_(
                Salary.employee_id == Employee.id,
                Salary.year == year,
                Salary.month == month
            ),
            isouter=True  # 使用外连接以确保能获取到所有部门数据
        ).group_by(
            Department.name,
            DepartmentBudget.total_budget,
            DepartmentBudget.salary_budget,
            DepartmentBudget.overtime_budget,
            DepartmentBudget.bonus_budget,
            DepartmentBudget.other_budget
        ).all()  # 移除 having 条件，获取所有数据
        
        # 转换为DataFrame，只包含异常数据
        df_data = []
        for row in exception_data:
            actual_expense = float(row[6]) if row[6] else 0
            total_budget = float(row[1])
            execution_rate = (actual_expense/total_budget*100) if total_budget > 0 else 0
            
            # 只添加执行率超过阈值的数据
            if execution_rate >= threshold:
                df_data.append({
                    '部门': row[0],
                    '总预算': total_budget,
                    '实际支出': actual_expense,
                    '预算执行率': f"{execution_rate:.1f}%",
                    '基本工资预算': float(row[2]),
                    '实际基本工资': float(row[7]) if row[7] else 0,
                    '工资执行率': f"{(float(row[7])/float(row[2])*100 if row[7] else 0):.1f}%",
                    '加班费预算': float(row[3]),
                    '实际加班费': float(row[8]) if row[8] else 0,
                    '加班费执行率': f"{(float(row[8])/float(row[3])*100 if row[8] and float(row[3]) > 0 else 0):.1f}%",
                    '奖金预算': float(row[4]),
                    '实际奖金': float(row[9]) if row[9] else 0,
                    '奖金执行率': f"{(float(row[9])/float(row[4])*100 if row[9] and float(row[4]) > 0 else 0):.1f}%",
                    '其他预算': float(row[5]),
                    '员工人数': row[10],
                    '人均支出': f"{actual_expense/row[10]:.2f}" if row[10] > 0 else "0.00",
                    '异常说明': get_exception_description(row)
                })
        
        # 创建DataFrame
        df = pd.DataFrame(df_data)
        
        # 如果没有异常数据，添加一行说明
        if df.empty:
            df = pd.DataFrame([{
                '部门': '无异常数据',
                '总预算': 0,
                '实际支出': 0,
                '预算执行率': '0.0%',
                '异常说明': f'所有部门预算执行率均低于{threshold}%'
            }])
        
        # 创建Excel文件
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='预算异常报告', index=False)
            
            # 获取工作表
            worksheet = writer.sheets['预算异常报告']
            
            # 设置列宽
            for idx, col in enumerate(df.columns):
                worksheet.column_dimensions[chr(65 + idx)].width = 15
                
            # 设置数字格式
            if not df.empty and '总预算' in df.columns:  # 确保有数据且包含相关列
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row[1:15:2]:  # 金额列
                        cell.number_format = '#,##0.00'
                
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'budget_exceptions_{year}_{month}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def get_exception_description(row):
    """生成异常说明"""
    exceptions = []
    
    # 检查总预算
    total_rate = (float(row[6])/float(row[1])*100)
    if total_rate >= 100:
        exceptions.append(f"总预算超支{(total_rate-100):.1f}%")
    elif total_rate >= 90:
        exceptions.append(f"总预算接近上限")
        
    # 检查工资预算
    salary_rate = (float(row[7])/float(row[2])*100)
    if salary_rate >= 100:
        exceptions.append(f"工资预算超支{(salary_rate-100):.1f}%")
        
    # 检查加班费预算
    if float(row[3]) > 0:
        ot_rate = (float(row[8])/float(row[3])*100)
        if ot_rate >= 100:
            exceptions.append(f"加班费超支{(ot_rate-100):.1f}%")
            
    # 检查奖金预算
    if float(row[4]) > 0:
        bonus_rate = (float(row[9])/float(row[4])*100)
        if bonus_rate >= 100:
            exceptions.append(f"奖金超支{(bonus_rate-100):.1f}%")
            
    return "；".join(exceptions) if exceptions else "无重大异常"